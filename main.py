import cv2
import os # untuk mebaca file
import numpy as np # untuk mebaca data array
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Path ke dataset dan model
DATASET_PATH = 'dataset'
MODEL_PATH = 'models'
# Path file Excel absensi
EXCEL_FILE = 'absensi.xlsx'


# Load model age & gender
age_model = cv2.dnn.readNetFromCaffe(
    os.path.join(MODEL_PATH, 'age_deploy.prototxt'),
    os.path.join(MODEL_PATH, 'age_net.caffemodel')
)
gender_model = cv2.dnn.readNetFromCaffe(
    os.path.join(MODEL_PATH, 'gender_deploy.prototxt'),
    os.path.join(MODEL_PATH, 'gender_net.caffemodel')
)

AGE_LIST = ['(0-2)', '(4-6)', '(8-12)', '(15-20)', '(25-32)', '(38-43)', '(48-53)', '(60-100)']
GENDER_LIST = ['Male', 'Female']

# Fungsi menyimpan ke Excel dengan mencegah absen dua kali
def simpan_ke_excel(nama, tanggal, waktu, path_file=EXCEL_FILE):
    now = datetime.now()
    tanggal = now.strftime("%Y-%m-%d")
    waktu = now.strftime("%H:%M:%S")

    if not os.path.exists(path_file):
        wb = Workbook()
        ws = wb.active
        ws.append(['Nama', 'Tanggal', 'Waktu'])
    else:
        wb = load_workbook(path_file)
        ws = wb.active

    # Cek apakah nama sudah absen di tanggal ini
    sudah_absen = any(
        row[0] == nama and row[1] == tanggal
        for row in ws.iter_rows(min_row=2, values_only=True)
    )

    if sudah_absen:
        print(f"[INFO] {nama} sudah absen pada {tanggal}.")
        return

    # Tambahkan baris baru
    ws.append([nama, tanggal, waktu])
    wb.save(path_file)
    print(f"[SUKSES] {nama} berhasil absen pada {tanggal} {waktu}")


# Inisialisasi model pengenalan wajah
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Fungsi untuk memuat dataset dengan label eksplisit
def get_images_and_labels(dataset_path):
    faces = []
    labels = []
    label_names = {}

    person_list = ['ridwan', 'fayyadh']  # urutan label: 0 = Andita, 1 = Shinta

    for label, person_name in enumerate(person_list):
        person_folder = os.path.join(dataset_path, person_name)
        if not os.path.isdir(person_folder):
            continue

        label_names[label] = person_name

        for filename in os.listdir(person_folder):
            if filename.lower().endswith(('.jpg', '.png')):
                image_path = os.path.join(person_folder, filename)
                image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
                if image is None:
                    print(f"Tidak bisa membaca gambar: {image_path}")
                    continue
                faces.append(image)
                labels.append(label)

    return faces, np.array(labels), label_names

# Load dataset
faces, labels, label_names = get_images_and_labels(DATASET_PATH)

# Training model
recognizer.train(faces, labels)

# Load classifier wajah
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

# Mulai webcam
cap = cv2.VideoCapture(0)
print("Tekan 'q' untuk keluar")

sudah_absen_nama = set()  # Mencegah spam dalam 1 sesi


while True:
    ret, frame = cap.read()
    if not ret:
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces_detected = face_cascade.detectMultiScale(gray, scaleFactor=1.2, minNeighbors=5)

    for (x, y, w, h) in faces_detected:
        roi_gray = gray[y:y+h, x:x+w]
        roi_color = frame[y:y+h, x:x+w]

        # Face recognition
        try:
            label, confidence = recognizer.predict(roi_gray)
            name = label_names.get(label, "Unknown")
            color = (0, 255, 0) if confidence < 100 else (0, 0, 255)
            identity = f"{name} ({confidence:.0f})" if confidence < 100 else "Unknown"
            display_text = f"{name} ({confidence:.0f})" if confidence < 100 else "Unknown"
            
            if confidence < 100 and name not in sudah_absen_nama:
                simpan_ke_excel(name)
                sudah_absen_nama.add(name)

        except:
            display_text = "Unknown"
            color = (0, 0, 255)

        # Gender & age prediction (gunakan ROI warna)
        face_blob = cv2.dnn.blobFromImage(roi_color, 1.0, (227, 227), (78.4263377603, 87.7689143744, 114.895847746), swapRB=False)
        gender_model.setInput(face_blob)
        gender_pred = gender_model.forward()
        gender = GENDER_LIST[gender_pred[0].argmax()]

        age_model.setInput(face_blob)
        age_pred = age_model.forward()
        age = AGE_LIST[age_pred[0].argmax()]

        label_text = f"{identity}, {gender}, {age}"

        # Tampilkan hasil
        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
        cv2.putText(frame, label_text, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

    cv2.imshow("Face Recognition + Age & Gender", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()